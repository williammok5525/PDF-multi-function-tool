"""
PDF格式转换功能
包括：PDF与图片互转、PDF转Word/Excel/Text、Office转PDF、HTML转PDF
"""

from pdf2image import convert_from_path
from PIL import Image
import fitz
import img2pdf
from pdf2docx import Converter
import pdfplumber
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from docx import Document
from openpyxl import Workbook
import os
import subprocess
from typing import List

class FormatConverter:
    """格式转换类"""
    
    @staticmethod
    def pdf_to_images(input_path: str, output_folder: str,
                     image_format: str = 'PNG', dpi: int = 200,
                     progress_callback=None) -> List[str]:
        """
        PDF转图片
        
        Args:
            image_format: 图片格式 PNG/JPG/BMP/TIFF
            dpi: 分辨率
        
        Returns:
            生成的图片路径列表
        """
        try:
            os.makedirs(output_folder, exist_ok=True)
            
            # 转换为图片
            images = convert_from_path(
                input_path,
                dpi=dpi,
                fmt=image_format.lower()
            )
            
            output_files = []
            total = len(images)
            
            for i, image in enumerate(images):
                output_file = os.path.join(
                    output_folder,
                    f'page_{i + 1}.{image_format.lower()}'
                )
                
                # 保存图片
                if image_format.upper() == 'JPG' or image_format.upper() == 'JPEG':
                    # JPG需要转换为RGB
                    rgb_image = image.convert('RGB')
                    rgb_image.save(output_file, image_format)
                else:
                    image.save(output_file, image_format)
                
                output_files.append(output_file)
                
                if progress_callback:
                    progress_callback(int((i + 1) / total * 100))
            
            return output_files
        except Exception as e:
            raise Exception(f"PDF转图片失败: {str(e)}")
    
    @staticmethod
    def images_to_pdf(image_list: List[str], output_path: str,
                     progress_callback=None) -> bool:
        """
        图片转PDF（支持多张图片）
        """
        try:
            # 使用img2pdf库，效果更好
            with open(output_path, "wb") as f:
                f.write(img2pdf.convert(image_list))
            
            if progress_callback:
                progress_callback(100)
            
            return True
        except Exception as e:
            # 备用方案：使用PIL
            try:
                images = []
                for img_path in image_list:
                    img = Image.open(img_path)
                    if img.mode == 'RGBA':
                        img = img.convert('RGB')
                    images.append(img)
                
                if images:
                    images[0].save(
                        output_path,
                        save_all=True,
                        append_images=images[1:] if len(images) > 1 else []
                    )
                    return True
            except Exception as e2:
                raise Exception(f"图片转PDF失败: {str(e2)}")
    
    @staticmethod
    def pdf_to_word(input_path: str, output_path: str,
                   progress_callback=None) -> bool:
        """PDF转Word"""
        try:
            cv = Converter(input_path)
            cv.convert(output_path, start=0, end=None)
            cv.close()
            
            if progress_callback:
                progress_callback(100)
            
            return True
        except Exception as e:
            raise Exception(f"PDF转Word失败: {str(e)}")
    
    @staticmethod
    def pdf_to_excel(input_path: str, output_path: str,
                    progress_callback=None) -> bool:
        """
        PDF转Excel（提取表格）
        """
        try:
            workbook = Workbook()
            workbook.remove(workbook.active)  # 删除默认sheet
            
            with pdfplumber.open(input_path) as pdf:
                total = len(pdf.pages)
                
                for page_num, page in enumerate(pdf.pages):
                    tables = page.extract_tables()
                    
                    if tables:
                        # 为每个包含表格的页面创建sheet
                        sheet = workbook.create_sheet(f'Page_{page_num + 1}')
                        
                        row_offset = 1
                        for table in tables:
                            for row in table:
                                for col_idx, cell in enumerate(row):
                                    sheet.cell(
                                        row=row_offset,
                                        column=col_idx + 1,
                                        value=cell
                                    )
                                row_offset += 1
                            row_offset += 1  # 表格间空一行
                    
                    if progress_callback:
                        progress_callback(int((page_num + 1) / total * 100))
            
            workbook.save(output_path)
            return True
        except Exception as e:
            raise Exception(f"PDF转Excel失败: {str(e)}")
    
    @staticmethod
    def pdf_to_text(input_path: str, output_path: str,
                   progress_callback=None) -> bool:
        """PDF转文本"""
        try:
            doc = fitz.open(input_path)
            total = len(doc)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                for page_num in range(total):
                    page = doc[page_num]
                    text = page.get_text()
                    f.write(f"{'='*50}\n")
                    f.write(f"第 {page_num + 1} 页\n")
                    f.write(f"{'='*50}\n")
                    f.write(text)
                    f.write("\n\n")
                    
                    if progress_callback:
                        progress_callback(int((page_num + 1) / total * 100))
            
            doc.close()
            return True
        except Exception as e:
            raise Exception(f"PDF转文本失败: {str(e)}")
    
    @staticmethod
    def word_to_pdf(input_path: str, output_path: str) -> bool:
        """
        Word转PDF（需要Microsoft Word或LibreOffice）
        """
        try:
            # 方法1: 使用docx2pdf (Windows only, 需要Word)
            try:
                import docx2pdf
                docx2pdf.convert(input_path, output_path)
                return True
            except:
                pass
            
            # 方法2: 使用LibreOffice (跨平台)
            try:
                subprocess.run([
                    'soffice',
                    '--headless',
                    '--convert-to', 'pdf',
                    '--outdir', os.path.dirname(output_path),
                    input_path
                ], check=True, timeout=60)
                return True
            except:
                pass
            
            raise Exception("需要安装Microsoft Word或LibreOffice")
            
        except Exception as e:
            raise Exception(f"Word转PDF失败: {str(e)}")
    
    @staticmethod
    def excel_to_pdf(input_path: str, output_path: str) -> bool:
        """Excel转PDF（需要LibreOffice）"""
        try:
            subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', os.path.dirname(output_path),
                input_path
            ], check=True, timeout=60)
            return True
        except Exception as e:
            raise Exception(f"Excel转PDF失败: {str(e)}")
    
    @staticmethod
    def html_to_pdf(html_path: str, output_path: str) -> bool:
        """HTML转PDF"""
        try:
            # 使用pdfkit (需要wkhtmltopdf)
            try:
                import pdfkit
                pdfkit.from_file(html_path, output_path)
                return True
            except:
                pass
            
            # 备用方案：使用weasyprint
            try:
                from weasyprint import HTML
                HTML(filename=html_path).write_pdf(output_path)
                return True
            except:
                pass
            
            raise Exception("需要安装wkhtmltopdf或weasyprint")
            
        except Exception as e:
            raise Exception(f"HTML转PDF失败: {str(e)}")
    
    @staticmethod
    def text_to_pdf(input_path: str, output_path: str,
                   font_path: str = None) -> bool:
        """
        文本转PDF
        
        Args:
            font_path: 中文字体路径（如需支持中文）
        """
        try:
            c = canvas.Canvas(output_path, pagesize=A4)
            width, height = A4
            
            # 注册中文字体
            if font_path and os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('CustomFont', font_path))
                font_name = 'CustomFont'
            else:
                font_name = 'Helvetica'
            
            with open(input_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            y = height - 50
            line_height = 15
            
            c.setFont(font_name, 12)
            
            for line in lines:
                if y < 50:  # 需要新页面
                    c.showPage()
                    c.setFont(font_name, 12)
                    y = height - 50
                
                c.drawString(50, y, line.rstrip())
                y -= line_height
            
            c.save()
            return True
        except Exception as e:
            raise Exception(f"文本转PDF失败: {str(e)}")